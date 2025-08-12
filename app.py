from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
import csv
import re
import uuid
import math
import os
import json

from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session
from openpyxl import Workbook, load_workbook


app = Flask(__name__)
app.secret_key = "order-stock-secret-key"

# In-memory result store keyed by a one-time token
RESULT_STORE: Dict[str, Dict[str, bytes]] = {}
# Batch store maps a batch token to a list of item tokens
BATCH_STORE: Dict[str, List[str]] = {}
# Simple schedule store: token -> ISO date string
SCHEDULE_STORE: Dict[str, str] = {}
# Decisions store for Not have items: key(fish|pack|order) -> decision
DECISION_STORE: Dict[str, str] = {}


# ----- Data Models -----
@dataclass
class Dataset:
    name: str
    rows: List[Dict[str, Any]]
    sheet_names: List[str]


# ----- Helpers -----
NORMALIZE_MAP = {
    "fish name": ["fish name", "fish", "product", "product name", "name"],
    "packed size": ["packed size", "pack", "pack size", "size"],
    "total carton": [
        "total carton",
        "total_ctn",
        "total ctn",
        "ctn",
        "carton",
        "cartons",
        "qty",
        "quantity",
    ],
    "weight_mc": [
        "weight_mc",
        "weight mt",
        "weight_mt",
        "net_weigh",
        "net weight",
        "weight per mc",
        "mc_weight",
        "kg/ctn",
        "kg per ctn",
    ],
}


def load_excel(file_storage, preferred_sheet: Optional[str] = None) -> Dataset:
    # Reset pointer and load workbook
    try:
        file_storage.stream.seek(0)
    except Exception:  # noqa: BLE001
        pass
    wb = load_workbook(file_storage, data_only=True)
    sheet_name = preferred_sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        headers = []
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        record = {}
        empty = True
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val not in (None, ""):
                empty = False
            record[h] = val
        if not empty:
            rows.append(record)

    return Dataset(name=getattr(file_storage, "filename", "uploaded.xlsx"), rows=rows, sheet_names=wb.sheetnames)


def load_csv(file_storage) -> Dataset:
    try:
        file_storage.stream.seek(0)
    except Exception:  # noqa: BLE001
        pass
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:  # noqa: BLE001
        text = raw.decode("latin-1", errors="ignore")
    reader = csv.DictReader(text.splitlines())
    rows: List[Dict[str, Any]] = []
    for r in reader:
        # Drop None keys if any
        rows.append({(k or ""): v for k, v in r.items()})
    return Dataset(name=getattr(file_storage, "filename", "uploaded.csv"), rows=rows, sheet_names=["CSV"])


def load_tabular(file_storage) -> Dataset:
    filename = (getattr(file_storage, "filename", "") or "").lower()
    if filename.endswith(".csv"):
        return load_csv(file_storage)
    # default to excel
    return load_excel(file_storage)


def try_map_row(row: Dict[str, Any]) -> Dict[str, Any]:
    lower_map = {k.lower().strip(): k for k in row.keys()}
    normalized: Dict[str, Any] = {}
    for key, aliases in NORMALIZE_MAP.items():
        value = None
        for alias in aliases:
            if alias in lower_map:
                value = row.get(lower_map[alias])
                break
        normalized[key] = value
    return normalized


def normalize_text_val(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().upper().replace("  ", " ")


_PARENS_RE = re.compile(r"\([^\)]*\)")
_MULTISPACE_RE = re.compile(r"\s+")
_RANGE_RE = re.compile(r"(\d+)\s*[-–]\s*(\d+)")
_PUNCT_BREAK_RE = re.compile(r"[\./_,]+")
_NON_ALNUM_RE = re.compile(r"[^A-Z0-9]+")


def canonicalize_product(text: Any) -> str:
    s = normalize_text_val(text)
    s = _PARENS_RE.sub(" ", s)
    s = _RANGE_RE.sub(lambda m: f"{m.group(1)} {m.group(2)}", s)
    # Ignore all special characters
    s = _NON_ALNUM_RE.sub(" ", s)
    # remove unit/noise tokens
    stop = {
        "G", "GM", "GMS", "GRAM", "GRAMS", "PCS", "PC", "CTN", "CTNS", "GL", "GLAZE",
        "WITH", "PRINT", "BAG", "RIDER", "STICKER", "PACK", "SIZE", "KG",
        # common glaze percentages as numbers
        "5", "10", "15", "20", "25", "30", "35", "40",
    }
    tokens = [t for t in _MULTISPACE_RE.split(s) if t]
    filtered: List[str] = []
    for t in tokens:
        if t in stop:
            continue
        filtered.append(t)
    # Return a compact key without spaces so 'SILVER CARP' == 'SILVERCARP'
    return "".join(filtered)


def canonicalize_pack(text: Any) -> str:
    s = normalize_text_val(text)
    s = _RANGE_RE.sub(lambda m: f"{m.group(1)} {m.group(2)}", s)
    # Ignore all special characters
    s = _NON_ALNUM_RE.sub(" ", s)
    s = _MULTISPACE_RE.sub(" ", s)
    # Return compact form (no spaces) to ignore spacing differences
    return s.replace(" ", "").strip()


_MASS_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(KG|KGS|G|GM|GRAM|GRAMS)")
_MULT_X_RE = re.compile(r"X\s*(\d+)")


def parse_kg_per_carton(text: Any) -> float:
    """Extract approximate kg per carton from a pack description, e.g. '1 KG X 10 BAG/CTN' -> 10.0.
    Returns 0.0 if cannot parse.
    """
    s = normalize_text_val(text)
    if not s:
        return 0.0
    s = s.replace("×", "X")
    # Combine first mass block with a nearby multiplier if present
    match = _MASS_RE.search(s)
    if not match:
        return 0.0
    value = float(match.group(1))
    unit = match.group(2)
    kg = value if unit.startswith("K") else value / 1000.0
    # Look ahead for a multiplier within the next ~20 chars
    tail = s[match.end():match.end() + 30]
    m2 = _MULT_X_RE.search(tail)
    mult = float(m2.group(1)) if m2 else 1.0
    return kg * mult


def to_int(val: Any) -> int:
    try:
        if val is None or val == "":
            return 0
        return int(float(val))
    except Exception:  # noqa: BLE001
        return 0


def compute_matches(stock_rows: List[Dict[str, Any]], order_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # Build stock lookups
    stock_by_prod_pack: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stock_by_prod: Dict[str, Dict[str, Any]] = {}

    for r in stock_rows:
        nr = try_map_row(r)
        prod_key = canonicalize_product(nr.get("fish name"))
        pack_key = canonicalize_pack(nr.get("packed size"))
        qty = to_int(nr.get("total carton"))
        kg_per_ctn = parse_kg_per_carton(nr.get("packed size"))
        if not prod_key and not pack_key:
            continue
        if prod_key:
            agg = stock_by_prod.get(prod_key, {"ctn": 0, "kg_per_ctn": kg_per_ctn})
            agg["ctn"] += qty
            # prefer non-zero kg_per_ctn when available
            if agg.get("kg_per_ctn", 0) == 0 and kg_per_ctn:
                agg["kg_per_ctn"] = kg_per_ctn
            stock_by_prod[prod_key] = agg
        key = (prod_key, pack_key)
        agg2 = stock_by_prod_pack.get(key, {"ctn": 0, "kg_per_ctn": kg_per_ctn})
        agg2["ctn"] += qty
        if agg2.get("kg_per_ctn", 0) == 0 and kg_per_ctn:
            agg2["kg_per_ctn"] = kg_per_ctn
        stock_by_prod_pack[key] = agg2

    results: List[Dict[str, Any]] = []
    for r in order_rows:
        nr = try_map_row(r)
        fish_text = nr.get("fish name")
        pack_text = nr.get("packed size")
        prod_key = canonicalize_product(fish_text)
        pack_key = canonicalize_pack(pack_text)
        order_qty = to_int(nr.get("total carton"))
        weight_mc = nr.get("weight_mc")
        try:
            order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
        except Exception:  # noqa: BLE001
            order_kg_per_ctn = parse_kg_per_carton(weight_mc)

        matched_by = ""
        stock_qty = 0
        # Prefer exact product+pack match when pack is present
        if pack_key and (prod_key, pack_key) in stock_by_prod_pack:
            stock_info = stock_by_prod_pack.get((prod_key, pack_key), {"ctn": 0, "kg_per_ctn": 0.0})
            stock_qty = int(stock_info.get("ctn", 0))
            stock_pack_kg = float(stock_info.get("kg_per_ctn", 0.0))
            matched_by = "product+pack"
        elif prod_key in stock_by_prod:
            stock_info = stock_by_prod.get(prod_key, {"ctn": 0, "kg_per_ctn": 0.0})
            stock_qty = int(stock_info.get("ctn", 0))
            stock_pack_kg = float(stock_info.get("kg_per_ctn", 0.0))
            matched_by = "product"
        else:
            stock_pack_kg = 0.0

        # Weight-aware calculations
        effective_stock_ctn = stock_qty
        mc_to_give = 0
        if order_kg_per_ctn and stock_pack_kg:
            total_stock_kg = stock_qty * stock_pack_kg
            # how many order-equivalent cartons can stock satisfy
            effective_stock_ctn = int(total_stock_kg // max(order_kg_per_ctn, 0.0001))
            # MC to pick from stock to satisfy order requirement
            required_kg = order_qty * order_kg_per_ctn
            mc_needed = math.ceil(required_kg / max(stock_pack_kg, 0.0001))
            mc_to_give = min(stock_qty, mc_needed)
        else:
            mc_to_give = min(stock_qty, order_qty)

        if effective_stock_ctn <= 0:
            status = "Not have"
        elif effective_stock_ctn < order_qty:
            status = "Not Full"
        else:
            status = "Full"

        fulfilled_ctn = min(order_qty, effective_stock_ctn)
        # Balance stock after giving the computed MC from stock
        balance_after_order = max(stock_qty - mc_to_give, 0)

        result = {
            "fish name": fish_text,
            "packed size": pack_text,
            "order_carton": order_qty,
            "stock_carton": stock_qty,  # raw MC from stock file
            "can_fulfill_carton": fulfilled_ctn,
            "shortfall": max(order_qty - effective_stock_ctn, 0),
            "status": status,
            "matched_by": matched_by,
            "order_kg_per_ctn": round(order_kg_per_ctn, 3) if order_kg_per_ctn else 0,
            "stock_kg_per_ctn": round(stock_pack_kg, 3) if stock_pack_kg else 0,
            "balance_stock_carton": balance_after_order,
            "mc_to_give": mc_to_give,
            "required_kg": round(order_qty * (order_kg_per_ctn or 0), 3),
        }
        results.append(result)

    return results


def rows_to_excel_bytes(rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()

    headers = [
        "Fish Name",
        "Packed Size",
        "Order CTN",
        "Stock CTN",
        "Can Fulfill",
        "Shortfall",
        "Status",
        "Order KG/CTN",
        "Stock KG/CTN",
        "Balance Stock CTN",
    ]

    def write_sheet(ws, data_rows: List[Dict[str, Any]]):
        ws.append(headers)
        for r in data_rows:
            ws.append([
                r.get("fish name", ""),
                r.get("packed size", ""),
                r.get("order_carton", 0),
                r.get("stock_carton", 0),
                r.get("can_fulfill_carton", 0),
                r.get("shortfall", 0),
                r.get("status", ""),
                r.get("order_kg_per_ctn", 0),
                r.get("stock_kg_per_ctn", 0),
                r.get("balance_stock_carton", 0),
            ])
        ws.freeze_panes = "A2"
        widths = [35, 18, 12, 12, 14, 12, 12, 14, 14, 18]
        for idx, width in enumerate(widths, start=1):
            col = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col].width = width

    # Create sheets: All, Full, Not Full, Not have
    ws_all = wb.active
    ws_all.title = "All"
    write_sheet(ws_all, rows)

    ws_full = wb.create_sheet(title="Full")
    write_sheet(ws_full, [r for r in rows if r.get("status") == "Full"])

    ws_nf = wb.create_sheet(title="Not Full")
    write_sheet(ws_nf, [r for r in rows if r.get("status") == "Not Full"])

    ws_nh = wb.create_sheet(title="Not have")
    write_sheet(ws_nh, [r for r in rows if r.get("status") == "Not have"])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def rows_to_pdf_bytes(rows: List[Dict[str, Any]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Order Availability Result", styles["Heading2"])
    story.append(title)
    story.append(Spacer(1, 10))

    headers = [
        "Fish Name",
        "Packed Size",
        "Order CTN",
        "Stock CTN",
        "Order KG/CTN",
        "Stock KG/CTN",
        "Balance Stock CTN",
        "Can Fulfill",
        "Shortfall",
        "Status",
    ]
    data = [headers]
    for r in rows:
        data.append([
            r.get("fish name", ""),
            r.get("packed size", ""),
            r.get("order_carton", 0),
            r.get("stock_carton", 0),
            r.get("order_kg_per_ctn", 0),
            r.get("stock_kg_per_ctn", 0),
            r.get("balance_stock_carton", 0),
            r.get("can_fulfill_carton", 0),
            r.get("shortfall", 0),
            r.get("status", ""),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fcfcfc")]),
    ]))
    story.append(table)

    doc.build(story)
    return buffer.getvalue()


# ----- Routes -----
@app.get("/")
def index():
    return render_template("index.html")


@app.post("/process")
def process():
    stock_file = request.files.get("stock_file")
    order_file = request.files.get("order_file")

    if not stock_file or not order_file:
        flash("Please upload both Stock and Order Excel files.", "error")
        return redirect(url_for("index"))

    try:
        stock_ds = load_tabular(stock_file)
        order_ds = load_tabular(order_file)
        result_rows = compute_matches(stock_ds.rows, order_ds.rows)
    except Exception as exc:  # noqa: BLE001
        flash(f"Failed to process files: {exc}", "error")
        return redirect(url_for("index"))

    def sum_required_kg(rows: List[Dict[str, Any]]) -> float:
        total = 0.0
        for r in rows:
            try:
                total += float(r.get("required_kg", 0) or 0)
            except Exception:  # noqa: BLE001
                pass
        return round(total, 3)

    summary = {
        "total_items": int(len(result_rows)),
        "full": int(sum(1 for r in result_rows if r["status"] == "Full")),
        "not_full": int(sum(1 for r in result_rows if r["status"] == "Not Full")),
        "not_have": int(sum(1 for r in result_rows if r["status"] == "Not have")),
        "total_kg_all": sum_required_kg(result_rows),
        "total_kg_full": sum_required_kg([r for r in result_rows if r["status"] == "Full"]),
        "total_kg_not_full": sum_required_kg([r for r in result_rows if r["status"] == "Not Full"]),
        "total_kg_not_have": sum_required_kg([r for r in result_rows if r["status"] == "Not have"]),
    }

    # store result in memory for download via token
    excel_bytes = rows_to_excel_bytes(result_rows)
    pdf_bytes = rows_to_pdf_bytes(result_rows)
    token = uuid.uuid4().hex
    order_basename = os.path.splitext(order_ds.name or "order")[0]
    RESULT_STORE[token] = {
        "excel": excel_bytes,
        "pdf": pdf_bytes,
        "excel_name": f"{order_basename} Calculation.xlsx",
        "pdf_name": f"{order_basename} Calculation.pdf",
        # Non-bytes metadata for rendering view routes
        "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
        "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
        "stock_name": stock_ds.name.encode("utf-8"),
        "order_name": order_ds.name.encode("utf-8"),
    }

    table_records = result_rows
    return render_template(
        "result.html",
        summary=summary,
        records=table_records,
        stock_name=stock_ds.name,
        order_name=order_ds.name,
        download_token=token,
    )


@app.get("/batch")
def batch_index():
    return render_template("batch.html")


@app.post("/process-batch")
def process_batch():
    stock_file = request.files.get("stock_file")
    order_files = request.files.getlist("order_files")
    if not stock_file or not order_files:
        flash("Please upload one Stock file and up to 32 Order files.", "error")
        return redirect(url_for("batch_index"))
    if len(order_files) > 32:
        flash("You can upload at most 32 order files.", "error")
        return redirect(url_for("batch_index"))

    try:
        stock_ds = load_tabular(stock_file)
    except Exception as exc:  # noqa: BLE001
        flash(f"Failed to read stock file: {exc}", "error")
        return redirect(url_for("batch_index"))

    batch_token = uuid.uuid4().hex
    token_list: List[str] = []

    # Accumulate unique Not have fish across all orders in this batch
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for of in order_files:
        try:
            order_ds = load_tabular(of)
            result_rows = compute_matches(stock_ds.rows, order_ds.rows)
            # Build summary like in single process
            def sum_required_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        total += float(r.get("required_kg", 0) or 0)
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            # Compute ready kg from stock present for this order (sum over rows of stock_ctn * stock_kg_per_ctn)
            def sum_ready_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        total += float(r.get("stock_carton", 0) or 0) * float(r.get("stock_kg_per_ctn", 0) or 0)
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            summary = {
                "total_items": int(len(result_rows)),
                "full": int(sum(1 for r in result_rows if r["status"] == "Full")),
                "not_full": int(sum(1 for r in result_rows if r["status"] == "Not Full")),
                "not_have": int(sum(1 for r in result_rows if r["status"] == "Not have")),
                "total_kg_all": sum_required_kg(result_rows),
                "total_kg_full": sum_required_kg([r for r in result_rows if r["status"] == "Full"]),
                "total_kg_not_full": sum_required_kg([r for r in result_rows if r["status"] == "Not Full"]),
                "total_kg_not_have": sum_required_kg([r for r in result_rows if r["status"] == "Not have"]),
                "ready_kg": sum_ready_kg(result_rows),
            }

            # Build per-order Not have fish, aggregated by fish+pack with needed kg
            for row in result_rows:
                if row.get("status") == "Not have":
                    fish_name = str(row.get("fish name", ""))
                    pack_text = str(row.get("packed size", ""))
                    needed_kg = float(row.get("required_kg", 0) or 0)
                    key = (fish_name, pack_text)
                    group = fish_groups.get(key)
                    if not group:
                        group = {
                            "fish_name": fish_name,
                            "packed_size": pack_text,
                            "total_needed_kg": 0.0,
                            "orders": [],
                        }
                        fish_groups[key] = group
                    group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                    # attach decision if already set
                    decision_key = f"{fish_name}|{pack_text}|{order_ds.name}"
                    decision = DECISION_STORE.get(decision_key)
                    group["orders"].append({
                        "order_name": order_ds.name,
                        "needed_kg": round(needed_kg, 3),
                        "decision": decision,
                    })

            excel_bytes = rows_to_excel_bytes(result_rows)
            pdf_bytes = rows_to_pdf_bytes(result_rows)
            token = uuid.uuid4().hex
            order_basename = os.path.splitext(order_ds.name or "order")[0]
            RESULT_STORE[token] = {
                "excel": excel_bytes,
                "pdf": pdf_bytes,
                "excel_name": f"{order_basename} Calculation.xlsx",
                "pdf_name": f"{order_basename} Calculation.pdf",
                "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
                "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
                "stock_name": stock_ds.name.encode("utf-8"),
                "order_name": order_ds.name.encode("utf-8"),
            }
            token_list.append(token)
        except Exception as exc:  # noqa: BLE001
            flash(f"Failed to process order file {getattr(of,'filename','unknown')}: {exc}", "error")

    BATCH_STORE[batch_token] = token_list

    # Build summary items
    items = []
    for t in token_list:
        entry = RESULT_STORE.get(t, {})
        try:
            # Decode stored summary/metadata
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            # If ready_kg missing (older entry), compute from rows
            if "ready_kg" not in summary:
                try:
                    rows_tmp = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
                    ready = 0.0
                    for rr in rows_tmp:
                        ready += float(rr.get("stock_carton", 0) or 0) * float(rr.get("stock_kg_per_ctn", 0) or 0)
                    summary["ready_kg"] = round(ready, 3)
                except Exception:
                    summary["ready_kg"] = 0.0

            items.append({
                "token": t,
                "order_name": order_name,
                "summary": summary,
                "scheduled_on": SCHEDULE_STORE.get(t),
            })
        except Exception:
            continue

    # Build chart data
    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]

    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

    events = []
    for it in items:
        if it.get("scheduled_on"):
            events.append({
                "title": it["order_name"],
                "start": it["scheduled_on"],
                "url": url_for('view_result', token=it["token"]),
                "extendedProps": {"token": it["token"]},
            })

    # Turn fish groups dict into a list for template
    fish_groups_list = [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]

    fish_total_kg = round(sum(g["total_needed_kg"] for g in fish_groups_list), 3)

    # Recommendations: sort orders by number of Full items desc; include ready_kg
    recommendations = sorted(
        (
            {
                "order_name": it["order_name"],
                "full": it["summary"].get("full", 0),
                "ready_kg": it["summary"].get("ready_kg", 0),
                "token": it["token"],
                "scheduled": bool(it.get("scheduled_on")),
            }
            for it in items
        ),
        key=lambda x: (x["full"], x["ready_kg"]),
        reverse=True,
    )

    return render_template(
        "summary.html",
        batch_token=batch_token,
        stock_name=stock_ds.name,
        items=items,
        fish_groups=fish_groups_list,
        fish_total_kg=fish_total_kg,
        recommendations=recommendations,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
        calendar_events=json.dumps(events),
    )


@app.get("/batch/<batch_token>")
def view_batch(batch_token: str):
    tokens = BATCH_STORE.get(batch_token)
    if not tokens:
        flash("Unknown or expired batch.", "error")
        return redirect(url_for("batch_index"))

    items = []
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stock_name = ""
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore") or stock_name
            # aggregate Not have by fish+pack with needed kg per order
            try:
                rows = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
                for row in rows:
                    if row.get("status") == "Not have":
                        fish_name = str(row.get("fish name", ""))
                        pack_text = str(row.get("packed size", ""))
                        needed_kg = float(row.get("required_kg", 0) or 0)
                        key = (fish_name, pack_text)
                        group = fish_groups.get(key)
                        if not group:
                            group = {
                                "fish_name": fish_name,
                                "packed_size": pack_text,
                                "total_needed_kg": 0.0,
                                "orders": [],
                            }
                            fish_groups[key] = group
                        group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                        decision_key = f"{fish_name}|{pack_text}|{order_name}"
                        decision = DECISION_STORE.get(decision_key)
                        group["orders"].append({
                            "order_name": order_name,
                            "needed_kg": round(needed_kg, 3),
                            "decision": decision,
                        })
            except Exception:
                pass

            items.append({
                "token": t,
                "order_name": order_name,
                "summary": summary,
                "scheduled_on": SCHEDULE_STORE.get(t),
            })
        except Exception:
            continue

    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]
    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)
    events = []
    for it in items:
        if it.get("scheduled_on"):
            events.append({
                "title": it["order_name"],
                "start": it["scheduled_on"],
                "url": url_for('view_result', token=it["token"]),
                "extendedProps": {"token": it["token"]},
            })

    fish_groups_list = [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]

    fish_total_kg = round(sum(g["total_needed_kg"] for g in fish_groups_list), 3)

    recommendations = sorted(
        (
            {
                "order_name": it["order_name"],
                "full": it["summary"].get("full", 0),
                "ready_kg": it["summary"].get("ready_kg", 0),
                "token": it["token"],
                "scheduled": bool(it.get("scheduled_on")),
            }
            for it in items
        ),
        key=lambda x: (x["full"], x["ready_kg"]),
        reverse=True,
    )

    return render_template(
        "summary.html",
        batch_token=batch_token,
        stock_name=stock_name,
        items=items,
        fish_groups=fish_groups_list,
        fish_total_kg=fish_total_kg,
        recommendations=recommendations,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
        calendar_events=json.dumps(events),
    )


@app.post("/set-decision")
def set_decision():
    fish_name = request.form.get('fish_name') or ''
    packed_size = request.form.get('packed_size') or ''
    order_name = request.form.get('order_name') or ''
    decision = request.form.get('decision') or ''
    batch_token = request.form.get('batch_token') or ''
    redirect_to = request.form.get('redirect_to') or ''
    key = f"{fish_name}|{packed_size}|{order_name}"
    if decision:
        DECISION_STORE[key] = decision
    else:
        DECISION_STORE.pop(key, None)
    if batch_token:
        if redirect_to == 'fish_buy':
            return redirect(url_for('fish_buy', batch_token=batch_token))
        # default: bring user back to summary with fish tab active (handled by JS init)
        return redirect(url_for('view_batch', batch_token=batch_token))
    return redirect(url_for('batch_index'))


def build_fish_groups_from_batch(batch_token: str) -> List[Dict[str, Any]]:
    tokens = BATCH_STORE.get(batch_token, [])
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            rows = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
            for row in rows:
                if row.get("status") == "Not have":
                    fish_name = str(row.get("fish name", ""))
                    pack_text = str(row.get("packed size", ""))
                    needed_kg = float(row.get("required_kg", 0) or 0)
                    key = (fish_name, pack_text)
                    group = fish_groups.get(key)
                    if not group:
                        group = {
                            "fish_name": fish_name,
                            "packed_size": pack_text,
                            "total_needed_kg": 0.0,
                            "orders": [],
                        }
                        fish_groups[key] = group
                    group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                    decision_key = f"{fish_name}|{pack_text}|{order_name}"
                    decision = DECISION_STORE.get(decision_key)
                    group["orders"].append({
                        "order_name": order_name,
                        "needed_kg": round(needed_kg, 3),
                        "decision": decision,
                    })
        except Exception:
            continue
    return [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]


@app.get("/fish-buy/<batch_token>")
def fish_buy(batch_token: str):
    # reuse batch view data
    tokens = BATCH_STORE.get(batch_token)
    if not tokens:
        flash("Unknown or expired batch.", "error")
        return redirect(url_for("batch_index"))

    # stock name from first token
    stock_name = ""
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        if stock_name:
            break

    items = []
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            items.append({"token": t, "order_name": order_name, "summary": summary})
        except Exception:
            continue

    fish_groups = build_fish_groups_from_batch(batch_token)

    # charts data (optional on this page)
    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]
    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

    return render_template(
        "fish_buy.html",
        batch_token=batch_token,
        stock_name=stock_name,
        items=items,
        fish_groups=fish_groups,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
    )


def fish_groups_to_excel_bytes(fish_groups: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fish Summary"
    ws.append(["Fish Name", "Packed Size", "Total Needed KG", "Orders Count"])
    for g in fish_groups:
        ws.append([g.get("fish_name", ""), g.get("packed_size", ""), g.get("total_needed_kg", 0), len(g.get("orders", []))])
    ws.freeze_panes = "A2"

    # Decisions sheet: only orders with a decision
    ws2 = wb.create_sheet("Decisions")
    ws2.append(["Fish Name", "Packed Size", "Order File", "Needed KG", "Decision"])
    for g in fish_groups:
        for o in g.get("orders", []):
            if o.get("decision"):
                ws2.append([g.get("fish_name", ""), g.get("packed_size", ""), o.get("order_name", ""), o.get("needed_kg", 0), o.get("decision")])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.get("/download/fish-excel")
def download_fish_excel():
    batch_token = request.args.get('batch', type=str)
    if not batch_token or batch_token not in BATCH_STORE:
        flash("Unknown batch.", "error")
        return redirect(url_for('batch_index'))
    fish_groups = build_fish_groups_from_batch(batch_token)
    raw = fish_groups_to_excel_bytes(fish_groups)
    return send_file(
        io.BytesIO(raw),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Fish Decisions - {batch_token}.xlsx",
    )


@app.post("/schedule-order")
def schedule_order():
    token = request.form.get('token')
    date = request.form.get('date')
    batch_token = request.form.get('batch_token')
    if token and date:
        SCHEDULE_STORE[token] = date
    if batch_token:
        return redirect(url_for('view_batch', batch_token=batch_token))
    return redirect(url_for('batch_index'))


@app.post("/unschedule-order")
def unschedule_order():
    token = request.form.get('token')
    batch_token = request.form.get('batch_token')
    if token:
        SCHEDULE_STORE.pop(token, None)
    # For fetch usage, return a simple OK
    return "OK"


@app.get("/result/<token>")
def view_result(token: str):
    entry = RESULT_STORE.get(token)
    if not entry:
        flash("Unknown or expired result token.", "error")
        return redirect(url_for("index"))
    try:
        rows = eval(entry.get("rows_json", b"[]"))  # noqa: S307
        summary = eval(entry.get("summary_json", b"{}"))  # noqa: S307
        stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
    except Exception:
        flash("Failed to load stored result.", "error")
        return redirect(url_for("index"))

    return render_template(
        "result.html",
        summary=summary,
        records=rows,
        stock_name=stock_name,
        order_name=order_name,
        download_token=token,
    )


@app.get("/download/excel")
def download_excel():
    token = request.args.get("token", type=str)
    raw = RESULT_STORE.get(token or "", {}).get("excel") if token else None
    if not raw:
        flash("No processed result to download yet.", "error")
        return redirect(url_for("index"))
    name = RESULT_STORE.get(token or "", {}).get("excel_name", "order_stock_result.xlsx")
    return send_file(
        io.BytesIO(raw),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=name,
    )


@app.get("/download/pdf")
def download_pdf():
    token = request.args.get("token", type=str)
    raw = RESULT_STORE.get(token or "", {}).get("pdf") if token else None
    if not raw:
        flash("No processed result to download yet.", "error")
        return redirect(url_for("index"))
    name = RESULT_STORE.get(token or "", {}).get("pdf_name", "order_stock_result.pdf")
    return send_file(
        io.BytesIO(raw),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=name,
    )


if __name__ == "__main__":
    app.run(debug=True)


